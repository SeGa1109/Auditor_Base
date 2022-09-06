import openpyxl

from Env import *

def WageCalcLay():
    head=['Emp Code','Name','F/S Name','Days Present','S1','S2','S3','Wage','OT','OT Wages','Incentive','Gross Wages','PF','ESI','Adv.','Canteen','Net Wages']
    headwidth=[12,25,25,5,5,5,5,10,5,10,10,15,10,10,10,10,20]
    TL_WCEXP=[[
    ms.Table(values=[], headings=head,
                justification='centre', enable_events=True,
                auto_size_columns=False,
                row_height=20,
                col_widths=headwidth,
                num_rows=100,
                font=fstyle,
                enable_click_events=True, key="TL_WC",
                hide_vertical_scroll=True
             )
    ]]
    layout=[
        [ms.Input(todatemy,font=fstyle,size=(10,1),key='wcdateinp'),ms.CalendarButton(" ",target='wcdateinp',format="%m-%Y"),
         ms.Button("Generate",key='wcgen',font=fstyle),ms.Sizer(400,0), ms.Text("Wages Report Generation",font=fstylehd ),ms.Sizer(400,0),
         ms.Button("Export",key='wcexp',font=fstyle,disabled=True),
         ms.Button("Mail", key='wcmail', font=fstyle, disabled=True)
         ],
       [ ms.Frame("Output",layout=[[ms.Column(TL_WCEXP,scrollable=True,size=(swi-50,shi-80),)]],font=fstyle,size=(swi-50,shi-80))
    ]
    ]
    return layout

def WageCalcFn(Menu,event,values):
    if event == 'wcgen':
        incentive_chk=int(ms.popup_get_text("Enter the number of days for incentive addition",no_titlebar=True,font=fstyle,location=(30,100)))
        incentive_amnt=float(ms.popup_get_text("Incentive Amount per day",no_titlebar=True,font=fstyle,location=(30,100)))
        data = attendance_fetch(values['wcdateinp'])
        wagedata = wage_fetch()
        globals()['wage_proc_data']=[]
        #print(data)
        for step in data:
            #try:
                #print(step)
                temp=[]
                for i in range (3):#EMP Details addition
                    temp.append(step[i])
                #print(step)
                chk = shiftcheck(step[0])
                print(chk)
                if chk == True:# if Emp is Shift resource
                    print(step)
                    wagetemp=wagedata.get(step[0])
                    print(wagetemp)
                    temp.append("NA")
                    S1,S2,S3,OT=0,0,0,0
                    CE=0.0
                    for i in range (4,len(step)):# Custom Shift Calc
                        try:
                            i=list(step[i].split(","))
                        except:
                            break
                        if i[0] == '1':
                            S1+=1
                        if i[0] == '2':
                            S2+=1
                        if i[0] == '3':
                            S3+=1
                        OT+=int(i[1]) # OT Addition
                        CE+=float(i[2]) # CE Addition
                    temp.append(S1)
                    temp.append(S2)
                    temp.append(S3)
                    wage = (S1*wagetemp[0])+(S2*wagetemp[1])+(S3*wagetemp[2]) # Wage Calc
                    temp.append(wage)
                    ot_wage = (OT / 8 * wagetemp[0])  # OT Calc
                    if step[3] == "yes":
                        incentive = 0.0
                    elif S1 + S2 + S3 >= incentive_chk:
                        incentive = incentive_amnt *(S1 + S2 + S3)
                    else:
                        incentive = 0.0
                else:
                    wagetemp = wagedata.get(step[0])
                    DP,OT=0,0
                    CE = 0.0
                    for i in range(4, len(step)):  # DP Calc
                        try:
                            i = list(step[i].split(","))
                        except:
                            break
                        if i[0] == 'P':
                            DP+=1
                        OT+=int(i[1]) # OT Addition
                        CE+=float(i[2]) # CE Addition
                    temp.append(DP)
                    temp.append("NA")
                    temp.append("NA")
                    temp.append("NA")
                    wage = DP*wagetemp
                    temp.append(wage)
                    print(OT,wagetemp)
                    ot_wage = (OT / 8 * wagetemp)
                    if step[3] == "yes":
                        incentive = 0.0
                    elif DP >= incentive_chk:
                        incentive = incentive_amnt *DP
                    else:
                        incentive = 0.0

                temp.append(OT)
                temp.append(ot_wage)
                temp.append(incentive)
                gross_wage=wage+ot_wage+incentive
                temp.append(gross_wage)
                PF = 0.0 if "SILTEMP" in step[0] else round(gross_wage*12/100,0)#PF Calculation
                temp.append(PF)
                ESI= 0.0 if "SILTEMP" in step[0] else round(gross_wage*0.75/100,0)#ESI Calculation
                temp.append(ESI)
                ADV=wageadvfetch(step[0],values['wcdateinp'])
                temp.append(ADV)
                temp.append(CE)#Canteen Expense Addition
                netwage=gross_wage-PF-ESI-ADV-CE
                temp.append(netwage)
                wage_proc_data.append(temp)
            #except Exception as e:
                #print(e)
        print(wage_proc_data)
        Menu['TL_WC'].update(values=wage_proc_data)
        Menu['wcexp'].update(disabled=False)
        Menu['wcmail'].update(disabled=False)

    if event == 'wcexp':
        xl=openpyxl.load_workbook(r"C:\Twink_06MA\Master_Files\Wage_Exp.xlsx")
        xl.active=xl['Wage_Calc']
        xlc=xl.active
        rowc=2
        colc=1
        wagedetails={}
        pfemp=[]
        nonpfemp=[]
        for step in wage_proc_data:
            wagedetails.update({step[0]:step[16]})
            colc=1
            for i in step:
                xlc.cell(row=rowc,column=colc).value=i
                colc+=1
            rowc+=1
            if "SILTEMP" in step[0]:
                nonpfemp.append(step)
            else:
                pfemp.append(step)
        nonpfempwage=[]
        for step in nonpfemp:
            mycursor.execute("select emp_code,employee_name,designation,bank_account_no,bank_name,"
                             "ifsc_code,branch from register where emp_code ='%s'"%step[0])
            db_data=list(sum(mycursor.fetchall(),()))
            db_data.append(step[16])
            nonpfempwage.append(db_data)
        pfempwage=[]
        for step in pfemp:
            mycursor.execute("select emp_code,employee_name,designation,bank_account_no,bank_name,"
                             "ifsc_code,branch from register where emp_code ='%s'"%step[0])
            db_data=list(sum(mycursor.fetchall(),()))
            db_data.append(step[16])
            pfempwage.append(db_data)

        print(pfempwage,nonpfempwage)
        xl.active=xl['PF']
        xlc=xl.active
        rowc=2
        colc=1
        for step in pfempwage:
            colc=1
            for i in step:
                xlc.cell(row=rowc, column=colc).value = i
                colc += 1
            rowc += 1

        xl.active = xl['NonPF']
        xlc = xl.active
        rowc = 2
        colc = 1
        for step in nonpfempwage:
            colc = 1
            for i in step:
                xlc.cell(row=rowc, column=colc).value = i
                colc += 1
            rowc += 1
        xl.save(r"C:\Twink_06MA\Master_Files\Wage_Exp_01.xlsx")
        os.system(r"C:\Twink_06MA\Master_Files\Wage_Exp_01.xlsx")

    if event == 'wcmail':
        maillist=popup_select(mailid_fetch(False,""), select_multiple=True)
        for i in maillist:
            mail_content = "PFA"
            sender_address = 'asta.sunilindustries@gmail.com'
            sender_pass = 'uxzgkfvkdzuxwpad'
            # Setup the MIME
            receiver_address = mailid_fetch(True,i)
            message = MIMEMultipart()
            message['From'] = sender_address
            message['To'] = receiver_address
            message['Subject'] = "Wage Calculation_Output"
            message.attach(MIMEText(mail_content, 'plain'))
            attach_file_name = r"C:\Twink_06MA\Master_Files\Wage_Exp_01.xlsx"
            attach_file = open(attach_file_name, 'rb')  # Open the file as binary mode
            payload = MIMEBase('application', 'octate-stream')
            payload.set_payload((attach_file).read())
            encoders.encode_base64(payload)  # encode the attachment
            # add payload header with filename
            payload.add_header('Content-Disposition ', 'attachment',
                               filename='Wage_Calc_Output.xlsx')
            message.attach(payload)
            # Create SMTP session for sending the mail
            session = smtplib.SMTP('smtp.gmail.com', 587)  # use gmail with port
            session.starttls()  # enable security
            session.login(sender_address, sender_pass)
            text = message.as_string()
            session.sendmail(sender_address, receiver_address, text)
            session.quit()
            print('Mail Sent')
        ms.popup_auto_close("Mail Successfully Sent", font=fstyle, no_titlebar=True)





